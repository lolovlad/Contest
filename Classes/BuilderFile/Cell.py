import openpyxl.utils
import openpyxl
import docx
from abc import ABC, abstractmethod


class Cell(ABC):
    def __init__(self, row, col, value):
        self._row = row
        self._col = col
        self.__value = value

    @abstractmethod
    def __str__(self):
        pass

    @property
    def value(self):
        return self.__value

    @property
    def row(self):
        return self._row

    @property
    def col(self):
        return self._col

    def _cell_from_string(self):
        return f"{openpyxl.utils.get_column_letter(self._col)}{self._row}"

    @abstractmethod
    def create_cell(self, sheet):
        pass


class SoloCell(Cell):
    def __init__(self, row, col, value):
        super().__init__(row, col, value)

    def __str__(self):
        return f"{self._cell_from_string()} {self.value}"

    def create_cell(self, sheet):
        if isinstance(sheet, openpyxl.workbook.workbook.Worksheet):
            sheet.cell(self._row, self._col).value = self.value
        else:
            cells = sheet.rows[self.row-1].cells
            cells[self.col-1].text = str(self.value)


class MergeCell(Cell):
    def __init__(self, row_start, col_start, row_end, col_end, value):
        super().__init__(row_start, col_start, value)
        self.__row_end = row_end
        self.__col_end = col_end

    def __str__(self):
        return f"{self._cell_from_string()} {self.value}"

    def __cell_merge_from_string(self):
        return f"{self._cell_from_string()}:{openpyxl.utils.get_column_letter(self.__col_end)}{self.__row_end}"

    def create_cell(self, sheet):
        if isinstance(sheet, openpyxl.workbook.workbook.Worksheet):
            sheet.merge_cells(self.__cell_merge_from_string())
            sheet.cell(self._row, self._col).value = self.value
        else:
            raise Exception("in docs not create")